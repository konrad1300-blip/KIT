#!/usr/bin/env python3
"""
Aplikacja KIT (Konstrukcja i Technologia) - System zarządzania pracami w dziale KiT
==============================================================================
Aplikacja webowa działająca w sieci lokalnej LAN z wbudowaną bazą danych SQLite.
Obsługuje dwa tryby pracy: użytkownik i administrator.

Główne funkcjonalności:
- System logowania z podziałem na role (admin, konstruktor, technolog, wdrożeniowiec)
- Zarządzanie zadaniami i przepływem pracy (workflow)
- Automatyczne tworzenie kroków procesu dla nowych zadań
- Przesyłanie i pobieranie plików (rysunki, BOM, instrukcje)
- Śledzenie postępu zadań poprzez pasek postępu
- Panel administratora do zarządzania użytkownikami
"""

# Import modułów standardowych
import os
import sqlite3
import io

# Import modułów Flask - framework webowy do obsługi żądań HTTP
from flask import Flask, render_template, request, redirect, url_for, session, flash, g, send_file, jsonify

# Import funkcji do bezpiecznego haszowania haseł
from werkzeug.security import generate_password_hash, check_password_hash

# Import dekoratora wraps do zachowania metadanych funkcji
from functools import wraps

# Import klasy datetime do operacji na datach i czasie
from datetime import datetime

# Import modułu os.path do operacji na ścieżkach plików
import os.path

# Import modułów do generowania raportów
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Inicjalizacja aplikacji Flask
app = Flask(__name__)

# Klucz sekretny używany do podpisywania ciasteczek sesji (bezpieczeństwo)
app.secret_key = 'kit-secret-key-2024'

# Katalog, w którym będą przechowywane wgrane pliki (rysunki, dokumenty)
app.config['UPLOAD_FOLDER'] = 'uploads'

# Maksymalny rozmiar przesyłanego pliku: 16MB (zabezpieczenie przed przeciążeniem)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max

# Nazwa pliku bazy danych SQLite (będzie utworzony w katalogu projektu)
DATABASE = 'kit.db'

def get_db():
    """
    Funkcja pomocnicza do obsługi połączenia z bazą danych SQLite.
    Wykorzystuje obiekt 'g' Flask do przechowywania połączenia w ramach jednego żądania.
    Dzięki temu wielokrotne wywołania get_db() w jednym żądaniu używają tego samego połączenia.
    
    Zwraca:
        sqlite3.Connection: Obiekt połączenia z bazą danych
    """
    # Sprawdź czy połączenie już istnieje w kontekście 'g' (obiekt Flask dla danego żądania)
    db = getattr(g, '_database', None)
    if db is None:
        # Utwórz nowe połączenie z bazą i zapisz w kontekście 'g'
        db = g._database = sqlite3.connect(DATABASE)
        # Ustaw row_factory na sqlite3.Row - pozwala to na dostęp do kolumn po nazwie (jak w słowniku)
        db.row_factory = sqlite3.Row
    return db


@app.teardown_appcontext
def close_connection(exception):
    """
    Funkcja wywoływana automatycznie przez Flask po zakończeniu każdego żądania.
    Zamyka połączenie z bazą danych, jeśli było otwarte.
    Dekorator @app.teardown_appcontext rejestruje tę funkcję jako callback.
    """
    # Pobierz połączenie z kontekstu 'g'
    db = getattr(g, '_database', None)
    if db is not None:
        # Zamknij połączenie z bazą danych
        db.close()

def init_db():
    """
    Inicjalizacja bazy danych - tworzenie wszystkich wymaganych tabel.
    Funkcja sprawdza czy tabele istnieją (CREATE TABLE IF NOT EXISTS)
    i tworzy domyślne role oraz konta administratorów.
    
    Tablice:
    - users: przechowuje dane użytkowników (login, hasło, rola)
    - tasks: przechowuje zadania/projekty
    - process_steps: kroki procesu dla każdego zadania
    - files: pliki powiązane z zadaniami
    - roles: lista dostępnych ról w systemie
    - skills: umiejętności/tagi
    - user_skills: relacja użytkownik-umiejętność
    """
    # Użyj kontekstu aplikacji Flask (wymagane do dostępu do bazy poza żądaniem)
    with app.app_context():
        db = get_db()
        cursor = db.cursor()
        
        # Tabela użytkowników - przechowuje dane logowania i role
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'user',
            full_name TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        ''')
        
        # Tabela zadań (projektów) - tworzone przez administratora
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS tasks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_name TEXT NOT NULL,
            description TEXT,
            status TEXT DEFAULT 'nowe',
            priority TEXT DEFAULT 'normal',
            due_date TIMESTAMP,
            admin_id INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            archived INTEGER DEFAULT 0,
            FOREIGN KEY (admin_id) REFERENCES users(id)
        )
        ''')
        
        # Migracja: dodanie kolumn jeśli nie istnieją
        columns = [row[1] for row in cursor.execute("PRAGMA table_info(tasks)").fetchall()]
        if 'priority' not in columns:
            cursor.execute("ALTER TABLE tasks ADD COLUMN priority TEXT DEFAULT 'normal'")
        if 'due_date' not in columns:
            cursor.execute("ALTER TABLE tasks ADD COLUMN due_date TIMESTAMP")
        if 'archived' not in columns:
            cursor.execute("ALTER TABLE tasks ADD COLUMN archived INTEGER DEFAULT 0")
        
        # Tabela kroków procesu - automatycznie generowana dla każdego zadania
        # Każdy krok ma przypisaną rolę, która może go wykonać
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS process_steps (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            task_id INTEGER NOT NULL,
            role_required TEXT NOT NULL,
            required_skill TEXT,
            description TEXT,
            status TEXT DEFAULT 'oczekujące',
            time_estimated REAL,
            time_actual REAL,
            completed_by INTEGER,
            completed_at TIMESTAMP,
            step_order INTEGER DEFAULT 0,
            FOREIGN KEY (task_id) REFERENCES tasks(id) ON DELETE CASCADE,
            FOREIGN KEY (completed_by) REFERENCES users(id)
        )
        ''')
        
        # Migracja: dodanie kolumny required_skill jeśli nie istnieje
        columns = [row[1] for row in cursor.execute("PRAGMA table_info(process_steps)").fetchall()]
        if 'required_skill' not in columns:
            cursor.execute("ALTER TABLE process_steps ADD COLUMN required_skill TEXT")
        
        # Tabela plików - przechowuje informacje o wgranych plikach
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS files (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            step_id INTEGER,
            task_id INTEGER NOT NULL,
            filename TEXT NOT NULL,
            file_path TEXT NOT NULL,
            file_type TEXT,
            uploaded_by INTEGER,
            uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (step_id) REFERENCES process_steps(id) ON DELETE CASCADE,
            FOREIGN KEY (task_id) REFERENCES tasks(id) ON DELETE CASCADE,
            FOREIGN KEY (uploaded_by) REFERENCES users(id)
        )
        ''')
        
        # Tabela ról - słownik dostępnych ról w systemie
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS roles (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL,
            display_name TEXT
        )
        ''')
        
        # Tabela umiejętności (skills/tags)
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS skills (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL,
            description TEXT,
            category TEXT
        )
        ''')
        
        # Tabela relacji użytkownik-umiejętność
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS user_skills (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            skill_id INTEGER NOT NULL,
            proficiency_level INTEGER DEFAULT 3,
            verified_at TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
            FOREIGN KEY (skill_id) REFERENCES skills(id),
            UNIQUE(user_id, skill_id)
        )
        ''')
        
        # Wstawienie domyślnych ról (jeśli jeszcze nie istnieją)
        cursor.execute("INSERT OR IGNORE INTO roles (name, display_name) VALUES ('admin', 'Administrator')")
        cursor.execute("INSERT OR IGNORE INTO roles (name, display_name) VALUES ('user', 'Użytkownik')")
        cursor.execute("INSERT OR IGNORE INTO roles (name, display_name) VALUES ('konstruktor', 'Konstruktor')")
        cursor.execute("INSERT OR IGNORE INTO roles (name, display_name) VALUES ('technolog', 'Technolog')")
        cursor.execute("INSERT OR IGNORE INTO roles (name, display_name) VALUES ('wdrozeniowiec', 'Wdrożeniowiec')")
        
        # Wstawienie przykładowych umiejętności
        default_skills = [
            ('AutoCAD', 'Projektowanie 2D/3D w AutoCad', 'CAD'),
            ('Inventor', 'Modelowanie 3D w Inventor', 'CAD'),
            ('SolidWorks', 'Modelowanie 3D w SolidWorks', 'CAD'),
            ('BOM', 'Przygotowanie BOM (Bill of Materials)', 'Technologia'),
            ('Wykończenie powierzchni', 'Oznaczanie wykończeń powierzchni', 'Technologia'),
            ('Lista procesów', 'Tworzenie list procesów z czasami', 'Technologia'),
            ('Wycena', 'Przygotowanie wyceny czasu i kosztów', 'Technologia'),
            ('Model 3D', 'Wykonanie modelu 3D', 'Modelowanie'),
            ('Instrukcje', 'Przygotowanie instrukcji krok-po-kroku', 'Dokumentacja'),
            ('Część mechaniczna', 'Doświadczenie w części mechanicznej', 'Mechanika'),
            ('Część hydrauliczna', 'Doświadczenie w hydraulice', 'Hydraulika'),
            ('Część elektryczna', 'Doświadczenie w elektryce', 'Elektryka'),
        ]
        for name, desc, cat in default_skills:
            cursor.execute("INSERT OR IGNORE INTO skills (name, description, category) VALUES (?, ?, ?)",
                          (name, desc, cat))
        
        # Sprawdź czy istnieje już jakiś administrator
        admin_exists = cursor.execute("SELECT id FROM users WHERE role='admin'").fetchone()
        if not admin_exists:
            # Utwórz domyślne konto administratora (login: admin, hasło: admin123)
            cursor.execute("INSERT INTO users (username, password_hash, role, full_name) VALUES (?, ?, ?, ?)",
                          ('admin', generate_password_hash('admin123'), 'admin', 'Administrator'))
         
        # Zatwierdź zmiany w bazie danych
        db.commit()

def login_required(f):
    """
    Dekorator sprawdzający czy użytkownik jest zalogowany.
    Jeśli w sesji nie ma 'user_id', przekierowuje na stronę logowania.
    Używa funkcji wraps z modułu functools, aby zachować metadane oryginalnej funkcji.
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Sprawdź czy użytkownik jest zalogowany (ma user_id w sesji)
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


def admin_required(f):
    """
    Dekorator sprawdzający czy zalogowany użytkownik ma rolę administratora.
    Jeśli nie - wyświetla komunikat o braku uprawnień i przekierowuje na stronę główną.
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Sprawdź czy użytkownik jest zalogowany i czy ma rolę 'admin'
        if 'user_id' not in session or session.get('role') != 'admin':
            flash('Brak uprawnień administratora')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function


@app.route('/')
@login_required
def index():
    """
    Główna strona aplikacji - dashboard użytkownika lub administratora.
    Po zalogowaniu użytkownik widzi swoje zadania, administrator widzi wszystkie zadania i użytkowników.
    """
    db = get_db()
    user_role = session.get('role')
    
    # Dla administratora - pokaż wszystkie zadania i listę użytkowników
    if user_role == 'admin':
        # Pobierz wszystkie zadania wraz z nazwą administratora, który je utworzył
        # Sortowanie: najpierw priorytet (bardzo pilne, pilne, normal), potem data utworzenia
        tasks = db.execute('''
            SELECT t.*, u.username as admin_name,
                   (SELECT COUNT(*) FROM process_steps WHERE task_id = t.id) as total_steps,
                   (SELECT COUNT(*) FROM process_steps WHERE task_id = t.id AND status = 'zakończone') as completed_steps
            FROM tasks t 
            LEFT JOIN users u ON t.admin_id = u.id 
            WHERE (t.archived = 0 OR t.archived IS NULL)
            ORDER BY 
                CASE t.priority 
                    WHEN 'bardzo pilne (natychmiast)' THEN 1
                    WHEN 'pilne (do końca dnia)' THEN 2
                    ELSE 3 
                END,
                t.created_at DESC
        ''').fetchall()
        # Pobierz wszystkich użytkowników
        users = db.execute('SELECT * FROM users ORDER BY username').fetchall()
        return render_template('admin_dashboard.html', tasks=tasks, users=users)
    else:
        # Dla zwykłego użytkownika - pokaż tylko zadania przypisane do jego roli
        # Oblicz też liczbę wszystkich kroków i ukończonych kroków dla każdego zadania
        # Sprawdzamy czy rola użytkownika znajduje się w kolumnie role_required (może zawierać wiele ról oddzielonych przecinkiem)
        # Sortowanie: najpierw priorytet, potem data utworzenia
        tasks = db.execute('''
            SELECT DISTINCT t.*, 
                   (SELECT COUNT(*) FROM process_steps WHERE task_id = t.id) as total_steps,
                   (SELECT COUNT(*) FROM process_steps WHERE task_id = t.id AND status = 'zakończone') as completed_steps
            FROM tasks t
            JOIN process_steps ps ON t.id = ps.task_id
            WHERE (',' || ps.role_required || ',' LIKE '%,' || ? || ',%') OR ? = 'admin'
            ORDER BY 
                CASE t.priority 
                    WHEN 'bardzo pilne (natychmiast)' THEN 1
                    WHEN 'pilne (do końca dnia)' THEN 2
                    ELSE 3 
                END,
                t.created_at DESC
        ''', (user_role, user_role)).fetchall()
        return render_template('user_dashboard.html', tasks=tasks)

@app.route('/login', methods=['GET', 'POST'])
def login():
    """
    Strona logowania - obsługa formularza logowania.
    GET: Wyświetla formularz logowania
    POST: Sprawdza dane logowania i tworzy sesję użytkownika
    """
    if request.method == 'POST':
        # Pobierz dane z formularza
        username = request.form['username']
        password = request.form['password']
        
        db = get_db()
        # Sprawdź czy użytkownik istnieje w bazie
        user = db.execute('SELECT * FROM users WHERE username = ?', (username,)).fetchone()
        
        # Sprawdź czy hasło jest poprawne (weryfikacja hasha)
        if user and check_password_hash(user['password_hash'], password):
            # Zapisz dane użytkownika w sesji
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['role'] = user['role']
            session['full_name'] = user['full_name']
            return redirect(url_for('index'))
        else:
            flash('Nieprawidłowy login lub hasło')
    
    # Wyświetl formularz logowania (GET) lub po błędzie logowania
    return render_template('login.html')


@app.route('/logout')
def logout():
    """
    Wylogowanie użytkownika - czyści wszystkie dane z sesji
    i przekierowuje na stronę logowania.
    """
    session.clear()
    return redirect(url_for('login'))


@app.route('/admin/users', methods=['GET', 'POST'])
@admin_required
def manage_users():
    """
    Panel zarządzania użytkownikami (tylko dla administratora).
    GET: Wyświetla listę użytkowników i formularz dodawania
    POST: Obsługuje akcje dodawania i usuwania użytkowników
    """
    db = get_db()
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        # Akcja dodawania nowego użytkownika
        if action == 'add':
            username = request.form['username']
            password = request.form['password']
            role = request.form['role']
            full_name = request.form.get('full_name', '')
            
            try:
                # Wstaw nowego użytkownika do bazy (hasło jest haszowane)
                db.execute('INSERT INTO users (username, password_hash, role, full_name) VALUES (?, ?, ?, ?)',
                         (username, generate_password_hash(password), role, full_name))
                db.commit()
                flash('Użytkownik dodany pomyślnie')
            except sqlite3.IntegrityError:
                # Użytkownik o tej nazwie już istnieje (naruszenie UNIQUE)
                flash('Użytkownik o tej nazwie już istnieje')
        
        # Akcja usuwania użytkownika
        elif action == 'delete':
            user_id = request.form['user_id']
            # Nie pozwól usunąć samego siebie
            if user_id != str(session['user_id']):
                # Sprawdź czy to nie jest administrator
                user_to_delete = db.execute('SELECT role FROM users WHERE id = ?', (user_id,)).fetchone()
                if user_to_delete and user_to_delete['role'] == 'admin':
                    flash('Nie można usunąć administratora')
                else:
                    db.execute('DELETE FROM users WHERE id = ?', (user_id,))
                    db.commit()
                    flash('Użytkownik usunięty')
            else:
                flash('Nie możesz usunąć samego siebie')
    
    # Pobierz listę wszystkich użytkowników posortowaną alfabetycznie
    users = db.execute('SELECT * FROM users ORDER BY username').fetchall()
    return render_template('manage_users.html', users=users)

@app.route('/admin/tasks/new', methods=['GET', 'POST'])
@admin_required
def new_task():
    """
    Tworzenie nowego zadania (tylko dla administratora).
    GET: Wyświetla formularz tworzenia zadania
    POST: Tworzy nowe zadanie i automatycznie generuje kroki procesu
    """
    db = get_db()
    
    if request.method == 'POST':
        project_name = request.form['project_name']
        description = request.form.get('description', '')
        priority = request.form.get('priority', 'normal')
        due_date = request.form.get('due_date') or None
        
        # Wstaw nowe zadanie do bazy (przypisane do obecnego administratora)
        cursor = db.execute('INSERT INTO tasks (project_name, description, priority, due_date, admin_id) VALUES (?, ?, ?, ?, ?)',
                           (project_name, description, priority, due_date, session['user_id']))
        # Pobierz ID nowo utworzonego zadania
        task_id = cursor.lastrowid
        
        # Pobierz szablony kroków z bazy (task_id = 0)
        template_steps = db.execute('SELECT * FROM process_steps WHERE task_id = 0 ORDER BY step_order, id').fetchall()
        
        # Jeśli nie ma szablonów, użyj domyślnych kroków
        if not template_steps:
            template_steps = [
                {'role_required': 'konstruktor', 'description': 'Przygotowanie rysunku ilustracji', 'step_order': 1},
                {'role_required': 'konstruktor', 'description': 'Wykonanie ilustracji', 'step_order': 2},
                {'role_required': 'konstruktor', 'description': 'Przygotowanie rysunku do kroju (DXF/DWG)', 'step_order': 3},
                {'role_required': 'technolog', 'description': 'Opracowanie BOM (Arkusz: Elementy)', 'step_order': 4},
                {'role_required': 'technolog', 'description': 'Lista procesów z czasami', 'step_order': 5},
                {'role_required': 'technolog', 'description': 'Przygotowanie wyceny', 'step_order': 6},
                {'role_required': 'technolog', 'description': 'Potwierdzenie w ZS i Microsoft SharePoint', 'step_order': 7},
                {'role_required': 'wdrozeniowiec', 'description': 'Wykonanie modelu 3D', 'step_order': 8},
                {'role_required': 'wdrozeniowiec', 'description': 'Przygotowanie instrukcji krok po kroku', 'step_order': 9}
            ]
        
        # Wstaw wszystkie kroki do bazy danych
        for i, step in enumerate(template_steps, 1):
            db.execute('INSERT INTO process_steps (task_id, role_required, description, step_order) VALUES (?, ?, ?, ?)',
                      (task_id, step['role_required'], step['description'], i))
        
        # Zatwierdź zmiany w bazie
        db.commit()
        flash('Zadanie utworzone pomyślnie')
        return redirect(url_for('index'))
    
    # Wyświetl formularz (metoda GET)
    return render_template('new_task.html')

@app.route('/uploads/<filename>')
@login_required
def serve_file(filename):
    """
    Serwowanie plików z katalogu uploads.
    Sprawdza czy plik istnieje i wysyła go do przeglądarki.
    Wymaga zalogowania (dekorator @login_required).
    """
    # Połącz nazwę pliku z katalogiem uploads
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if os.path.exists(filepath):
        # Wyślij plik do przeglądarki (możliwość otwarcia np. PDF w przeglądarce)
        return send_file(filepath)
    flash('Plik nie istnieje')
    return redirect(url_for('index'))


@app.route('/task/<int:task_id>/upload', methods=['POST'])
@login_required
def upload_file(task_id):
    """
    Wgrywanie plików dla konkretnego zadania.
    Zapisuje plik w katalogu uploads i dodaje rekord do bazy danych.
    Nazwa pliku zawiera ID zadania i timestamp, aby uniknąć kolizji nazw.
    """
    # Sprawdź czy w żądaniu jest plik
    if 'file' not in request.files:
        flash('Brak pliku')
        return redirect(url_for('view_task', task_id=task_id))
    
    file = request.files['file']
    # Sprawdź czy użytkownik wybrał plik
    if file.filename == '':
        flash('Nie wybrano pliku')
        return redirect(url_for('view_task', task_id=task_id))
    
    db = get_db()
    # Pobierz ID kroku (opcjonalne - plik może być ogólny dla zadania)
    step_id = request.form.get('step_id') or None
    
    # Wygeneruj unikalną nazwę pliku (zapobiega nadpisaniu plików o tej samej nazwie)
    filename = f"{task_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}"
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    
    # Zapisz plik na dysku
    file.save(filepath)
    
    # Dodaj informacje o pliku do bazy danych
    db.execute('INSERT INTO files (task_id, step_id, filename, file_path, file_type, uploaded_by) VALUES (?, ?, ?, ?, ?, ?)',
              (task_id, step_id, filename, filepath, file.content_type, session['user_id']))
    db.commit()
    
    flash('Plik wysłany')
    return redirect(url_for('view_task', task_id=task_id))


@app.route('/task/<int:task_id>')
@login_required
def view_task(task_id):
    """
    Wyświetlanie szczegółów zadania wraz z krokami procesu i plikami.
    Pobiera dane zadania, przypisane kroki oraz listę plików.
    """
    db = get_db()
    # Pobierz dane zadania
    task = db.execute('SELECT * FROM tasks WHERE id = ?', (task_id,)).fetchone()
    
    # Sprawdź czy zadanie istnieje
    if not task:
        flash('Zadanie nie istnieje')
        return redirect(url_for('index'))
    
    # Pobierz wszystkie kroki procesu dla tego zadania (posortowane według kolejności)
    steps = db.execute('SELECT * FROM process_steps WHERE task_id = ? ORDER BY step_order', (task_id,)).fetchall()
    
    # Pobierz wszystkie pliki przypisane do zadania wraz z nazwą użytkownika, który je wgrał
    files = db.execute('SELECT f.*, u.username FROM files f LEFT JOIN users u ON f.uploaded_by = u.id WHERE f.task_id = ?', (task_id,)).fetchall()
    
    # Pobierz umiejętności zalogowanego użytkownika (jeśli jest zalogowany)
    user_skill_names = []
    if session.get('user_id'):
        user_skills = db.execute('''
            SELECT s.name FROM user_skills us
            JOIN skills s ON us.skill_id = s.id
            WHERE us.user_id = ?
        ''', (session['user_id'],)).fetchall()
        user_skill_names = [s['name'] for s in user_skills]
    
    return render_template('view_task.html', task=task, steps=steps, files=files, user_skill_names=user_skill_names)


@app.route('/task/<int:task_id>/step/<int:step_id>/complete', methods=['POST'])
@login_required
def complete_step(task_id, step_id):
    """
    Kończenie kroku procesu przez użytkownika.
    Sprawdza czy zalogowany użytkownik ma odpowiednią rolę do wykonania kroku.
    Obsługuje wiele ról w jednym kroku (oddzielone przecinkiem, np. 'konstruktor,technolog').
    Po zakończeniu wszystkich kroków, zadanie otrzymuje status 'zakończone'.
    """
    db = get_db()
    user_role = session.get('role')
    
    # Pobierz dane kroku
    step = db.execute('SELECT * FROM process_steps WHERE id = ? AND task_id = ?', (step_id, task_id)).fetchone()
    
    # Sprawdź czy rola użytkownika znajduje się w liście dozwolonych ról dla tego kroku
    # Role są przechowywane jako string oddzielony przecinkami (np. 'konstruktor,technolog')
    allowed_roles = [r.strip() for r in step['role_required'].split(',')]
    if user_role not in allowed_roles:
        flash('To zadanie nie jest przypisane do Twojej roli')
        return redirect(url_for('view_task', task_id=task_id))
    
    # Sprawdź czy krok ma wymaganą umiejętność
    required_skill = step['required_skill'] if step['required_skill'] else None
    if required_skill:
        # Sprawdź czy użytkownik posiada tę umiejętność
        skill_check = db.execute('''
            SELECT us.id FROM user_skills us
            JOIN skills s ON us.skill_id = s.id
            WHERE us.user_id = ? AND s.name = ?
        ''', (session['user_id'], required_skill)).fetchone()
        
        if not skill_check:
            flash(f'Do wykonania tego kroku wymagana jest umiejętność: {required_skill}')
            return redirect(url_for('view_task', task_id=task_id))
    
    # Sprawdź czy krok nie został już ukończony
    if step['status'] == 'zakończone':
        flash('Ten krok został już zakończony')
        return redirect(url_for('view_task', task_id=task_id))
    
    # Pobierz czas wykonania wpisany przez użytkownika
    time_actual = request.form.get('time_actual')
    
    # Zaktualizuj krok jako zakończony
    db.execute('UPDATE process_steps SET status = ?, completed_by = ?, completed_at = ?, time_actual = ? WHERE id = ?',
             ('zakończone', session['user_id'], datetime.now(), time_actual, step_id))
    
    # Sprawdź status zadania
    all_steps = db.execute('SELECT COUNT(*) as total, SUM(CASE WHEN status = "zakończone" THEN 1 ELSE 0 END) as completed FROM process_steps WHERE task_id = ?', (task_id,)).fetchone()
    
    # Logika statusu:
    # - Jeśli wszystkie kroki ukończone -> "Zrobione" (zielony)
    # - Jeśli przynajmniej jeden krok ukończony -> "w toku" (pomarańczowy)
    # - W przeciwnym razie pozostaje "nowe" (szary)
    if all_steps['total'] == all_steps['completed'] and all_steps['total'] > 0:
        # Wszystkie kroki ukończone
        db.execute('UPDATE tasks SET status = ? WHERE id = ?', ('zakończone', task_id))
    elif all_steps['completed'] > 0:
        # Przynajmniej jeden krok ukończony, ale nie wszystkie
        db.execute('UPDATE tasks SET status = ? WHERE id = ?', ('w toku', task_id))
    # else: pozostaje "nowe"
    
    # Zatwierdź zmiany w bazie
    db.commit()
    flash('Krok zakończony')
    return redirect(url_for('view_task', task_id=task_id))


@app.route('/task/<int:task_id>/step/<int:step_id>/assign', methods=['POST'])
@admin_required
def assign_step_performer(task_id, step_id):
    """
    Przypisanie konkretnego użytkownika do wykonania kroku (admin).
    """
    db = get_db()
    user_id = request.form.get('user_id')
    
    if user_id:
        # Sprawdź czy krok istnieje
        step = db.execute('SELECT * FROM process_steps WHERE id = ? AND task_id = ?', 
                         (step_id, task_id)).fetchone()
        if not step:
            flash('Krok nie istnieje')
            return redirect(url_for('view_task', task_id=task_id))
        
        # Sprawdź czy użytkownik ma odpowiednią rolę
        user = db.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
        if not user:
            flash('Użytkownik nie istnieje')
            return redirect(url_for('view_task', task_id=task_id))
        
        # Sprawdź czy rola użytkownika jest dozwolona dla tego kroku
        allowed_roles = [r.strip() for r in step['role_required'].split(',')]
        if user['role'] not in allowed_roles and user['role'] != 'admin':
            flash(f'Użytkownik {user["username"]} nie ma odpowiedniej roli')
            return redirect(url_for('view_task', task_id=task_id))
        
        # Sprawdź czy wymagana umiejętność
        if step['required_skill']:
            skill_check = db.execute('''
                SELECT us.id FROM user_skills us
                JOIN skills s ON us.skill_id = s.id
                WHERE us.user_id = ? AND s.name = ?
            ''', (user_id, step['required_skill'])).fetchone()
            if not skill_check:
                flash(f'Użytkownik {user["username"]} nie posiada wymaganej umiejętności: {step["required_skill"]}')
                return redirect(url_for('view_task', task_id=task_id))
        
        # Przypisz użytkownika (możemy dodać kolumnę assigned_to jeśli potrzeba)
        # Na razie informacyjnie
        flash(f'Użytkownik {user["username"]} przypisany do kroku (sprawdź kompetencje)')
    else:
        flash('Nie wybrano użytkownika')
    
    return redirect(url_for('view_task', task_id=task_id))


@app.route('/reports', methods=['GET', 'POST'])
@login_required
def reports():
    """
    Strona raportów - generowanie raportów z wykonanych zadań.
    Obsługuje filtry: zakres dat, użytkownik (dla admina).
    Dla zwykłego użytkownika pokazuje tylko jego wykonane kroki.
    """
    db = get_db()
    user_role = session.get('role')
    user_id = session.get('user_id')
    
    # Pobierz listę użytkowników (dla administratora do filtrowania)
    users = []
    if user_role == 'admin':
        users = db.execute('SELECT * FROM users ORDER BY username').fetchall()
    
    # Parametry filtrowania
    date_from = request.form.get('date_from', '')
    date_to = request.form.get('date_to', '')
    filter_user_id = request.form.get('user_id', '')
    
    # Budowanie zapytania SQL z filtrami
    query = '''
        SELECT ps.*, 
               t.project_name,
               u.username as completed_by_name,
               u.full_name as completed_by_full_name
        FROM process_steps ps
        JOIN tasks t ON ps.task_id = t.id
        LEFT JOIN users u ON ps.completed_by = u.id
        WHERE ps.status = 'zakończone'
    '''
    params = []
    
    # Filtruj według użytkownika
    if user_role != 'admin':
        # Zwykły użytkownik widzi tylko swoje kroki
        query += ' AND ps.completed_by = ?'
        params.append(user_id)
    elif filter_user_id:
        # Admin może filtrować po konkretnym użytkowniku
        query += ' AND ps.completed_by = ?'
        params.append(filter_user_id)
    
    # Filtruj według zakresu dat
    if date_from:
        query += ' AND DATE(ps.completed_at) >= ?'
        params.append(date_from)
    if date_to:
        query += ' AND DATE(ps.completed_at) <= ?'
        params.append(date_to)
    
    query += ' ORDER BY ps.completed_at DESC'
    
    # Pobierz dane do raportu
    report_data = db.execute(query, params).fetchall()
    
    
    # Oblicz całkowity czas wykonania w Pythonie
    total_time = sum(float(row["time_actual"]) for row in report_data if row["time_actual"] is not None)
    
    return render_template('reports.html', 
                         report_data=report_data,
                         total_time=total_time,
                         users=users,
                         date_from=date_from,
                         date_to=date_to,
                         filter_user_id=filter_user_id)


@app.route('/reports/export/pdf', methods=['POST'])
@login_required
def export_pdf():
    """
    Eksport raportu do formatu PDF.
    Generuje plik PDF z wykonanymi krokami procesu.
    """
    db = get_db()
    user_role = session.get('role')
    user_id = session.get('user_id')
    
    # Parametry filtrowania
    date_from = request.form.get('date_from', '')
    date_to = request.form.get('date_to', '')
    filter_user_id = request.form.get('user_id', '')
    
    # Budowanie zapytania SQL z filtrami
    query = '''
        SELECT ps.*, 
               t.project_name,
               u.username as completed_by_name,
               u.full_name as completed_by_full_name
        FROM process_steps ps
        JOIN tasks t ON ps.task_id = t.id
        LEFT JOIN users u ON ps.completed_by = u.id
        WHERE ps.status = 'zakończone'
    '''
    params = []
    
    if user_role != 'admin':
        query += ' AND ps.completed_by = ?'
        params.append(user_id)
    elif filter_user_id:
        query += ' AND ps.completed_by = ?'
        params.append(filter_user_id)
    
    if date_from:
        query += ' AND DATE(ps.completed_at) >= ?'
        params.append(date_from)
    if date_to:
        query += ' AND DATE(ps.completed_at) <= ?'
        params.append(date_to)
    
    query += ' ORDER BY ps.completed_at DESC'
    
    report_data = db.execute(query, params).fetchall()
    
    # Tworzenie PDF
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", size=10)
    
    # Nagłówek
    pdf.set_font("Helvetica", 'B', 16)
    pdf.cell(0, 10, 'Raport wykonanych zadan', 0, 1, 'C')
    pdf.set_font("Helvetica", size=10)
    pdf.cell(0, 10, f'Wygenerowano: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', 0, 1, 'C')
    pdf.ln(10)
    
    # Filtry
    if date_from or date_to:
        pdf.cell(0, 10, f'Zakres dat: {date_from or "..."} - {date_to or "..."}', 0, 1)
    if filter_user_id and user_role == 'admin':
        user = db.execute('SELECT username FROM users WHERE id = ?', (filter_user_id,)).fetchone()
        if user:
            pdf.cell(0, 10, f'Uzytkownik: {user["username"]}', 0, 1)
    pdf.ln(5)
    
    # Tabela - nagłówki
    col_widths = [15, 45, 55, 35, 25, 25]
    headers = ['ID', 'Projekt', 'Opis kroku', 'Wykonane przez', 'Data', 'Czas (h)']
    
    pdf.set_fill_color(200, 220, 255)
    for i, header in enumerate(headers):
        pdf.cell(col_widths[i], 10, header, 1, 0, 'C', 1)
    pdf.ln()
    
    # Dane
    pdf.set_fill_color(255, 255, 255)
    for row in report_data:
        pdf.cell(col_widths[0], 10, str(row['id']), 1, 0, 'C')
        proj_name = row['project_name'][:20] if row['project_name'] else ''
        pdf.cell(col_widths[1], 10, proj_name, 1, 0)
        desc = row['description'][:25] if row['description'] else ''
        pdf.cell(col_widths[2], 10, desc, 1, 0)
        name = row['completed_by_full_name'] or row['completed_by_name'] or ''
        name = name[:15] if name else ''
        pdf.cell(col_widths[3], 10, name, 1, 0)
        completed_date = str(row['completed_at'])[:10] if row['completed_at'] else ''
        pdf.cell(col_widths[4], 10, completed_date, 1, 0, 'C')
        time_val = str(row['time_actual']) if row['time_actual'] else ''
        pdf.cell(col_widths[5], 10, time_val, 1, 1, 'C')
    
    # Podsumowanie
    pdf.ln(10)
    total_time = sum([float(row['time_actual']) if row['time_actual'] else 0 for row in report_data])
    pdf.cell(0, 10, f'Laczny czas: {total_time:.2f} h', 0, 1)
    pdf.cell(0, 10, f'Liczba krokow: {len(report_data)}', 0, 1)
    
    # Zapisz do bufora
    pdf_output = pdf.output(dest='S')
    buffer = io.BytesIO(pdf_output)
    buffer.seek(0)
    
    return send_file(buffer, as_attachment=True, download_name=f'raport_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf', mimetype='application/pdf')


@app.route('/reports/export/excel', methods=['POST'])
@login_required
def export_excel():
    """
    Eksport raportu do formatu Excel.
    Generuje plik Excel z wykonanymi krokami procesu.
    """
    db = get_db()
    user_role = session.get('role')
    user_id = session.get('user_id')
    
    # Parametry filtrowania
    date_from = request.form.get('date_from', '')
    date_to = request.form.get('date_to', '')
    filter_user_id = request.form.get('user_id', '')
    
    # Budowanie zapytania SQL z filtrami
    query = '''
        SELECT ps.*, 
               t.project_name,
               u.username as completed_by_name,
               u.full_name as completed_by_full_name
        FROM process_steps ps
        JOIN tasks t ON ps.task_id = t.id
        LEFT JOIN users u ON ps.completed_by = u.id
        WHERE ps.status = 'zakończone'
    '''
    params = []
    
    if user_role != 'admin':
        query += ' AND ps.completed_by = ?'
        params.append(user_id)
    elif filter_user_id:
        query += ' AND ps.completed_by = ?'
        params.append(filter_user_id)
    
    if date_from:
        query += ' AND DATE(ps.completed_at) >= ?'
        params.append(date_from)
    if date_to:
        query += ' AND DATE(ps.completed_at) <= ?'
        params.append(date_to)
    
    query += ' ORDER BY ps.completed_at DESC'
    
    report_data = db.execute(query, params).fetchall()
    
    # Tworzenie Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Raport"
    
    # Style nagłówka
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Nagłówki
    headers = ['ID', 'Projekt', 'Opis kroku', 'Rola', 'Wykonane przez', 'Data wykonania', 'Czas (h)', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Dane
    for row_idx, row in enumerate(report_data, 2):
        ws.cell(row=row_idx, column=1, value=row['id'])
        ws.cell(row=row_idx, column=2, value=row['project_name'])
        ws.cell(row=row_idx, column=3, value=row['description'])
        ws.cell(row=row_idx, column=4, value=row['role_required'])
        name = row['completed_by_full_name'] or row['completed_by_name'] or ''
        ws.cell(row=row_idx, column=5, value=name)
        ws.cell(row=row_idx, column=6, value=str(row['completed_at']) if row['completed_at'] else '')
        ws.cell(row=row_idx, column=7, value=float(row['time_actual']) if row['time_actual'] else 0)
        ws.cell(row=row_idx, column=8, value=row['status'])
    
    # Podsumowanie
    summary_row = len(report_data) + 3
    total_time = sum([float(row['time_actual']) if row['time_actual'] else 0 for row in report_data])
    ws.cell(row=summary_row, column=1, value="Łączny czas:")
    ws.cell(row=summary_row, column=2, value=f"{total_time:.2f} h")
    ws.cell(row=summary_row+1, column=1, value="Liczba kroków:")
    ws.cell(row=summary_row+1, column=2, value=len(report_data))
    
    # Szerokości kolumn
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    
    # Zapisz do bufora
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(buffer, as_attachment=True, 
                   download_name=f'raport_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
                   mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/admin/steps', methods=['GET', 'POST'])
@admin_required
def manage_steps():
    """
    Zarządzanie krokami procesu (tylko dla administratora).
    GET: Wyświetla listę kroków i formularz dodawania
    POST: Obsługuje akcje dodawania i usuwania kroków
    """
    db = get_db()
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        # Akcja dodawania nowego kroku
        if action == 'add':
            description = request.form['description']
            role_required = request.form['role_required']
            skill_id = request.form.get('required_skill', '')
            
            # Pobierz nazwę umiejętności jeśli podano
            required_skill = None
            if skill_id:
                skill = db.execute('SELECT name FROM skills WHERE id = ?', (skill_id,)).fetchone()
                if skill:
                    required_skill = skill['name']
            
            try:
                db.execute('INSERT INTO process_steps (task_id, role_required, required_skill, description, status, step_order) VALUES (?, ?, ?, ?, ?, ?)',
                          (0, role_required, required_skill, description, 'template', 0))  # task_id=0 oznacza szablon
                db.commit()
                flash('Krok procesu dodany pomyślnie')
            except Exception as e:
                flash(f'Błąd: {str(e)}')
        
        # Akcja usuwania kroku
        elif action == 'delete':
            step_id = request.form.get('step_id')
            if step_id:
                db.execute('DELETE FROM process_steps WHERE id = ? AND task_id = 0', (step_id,))
                db.commit()
                flash('Krok procesu usunięty')
    
    # Pobierz listę szablonów kroków (task_id = 0)
    steps = db.execute('SELECT * FROM process_steps WHERE task_id = 0 ORDER BY step_order, id').fetchall()
    
    # Pobierz wszystkie umiejętności (dla ustawiania required_skill)
    skills = db.execute('SELECT * FROM skills ORDER BY category, name').fetchall()
    
    return render_template('manage_steps.html', steps=steps, skills=skills)


@app.route('/admin/tasks/<int:task_id>/archive', methods=['POST'])
@admin_required
def archive_task(task_id):
    """
    Archiwizacja zadania (tylko dla administratora).
    Ustawia flagę archived na 1 dla danego zadania.
    """
    db = get_db()
    
    # Sprawdź czy zadanie istnieje
    task = db.execute('SELECT * FROM tasks WHERE id = ?', (task_id,)).fetchone()
    if not task:
        flash('Zadanie nie istnieje')
        return redirect(url_for('index'))
    
    # Archiwizuj zadanie
    db.execute('UPDATE tasks SET archived = 1 WHERE id = ?', (task_id,))
    db.commit()
    
    flash(f'Zadanie "{task["project_name"]}" zostało zarchiwizowane')
    return redirect(url_for('index'))


@app.route('/admin/tasks/<int:task_id>/delete', methods=['POST'])
@admin_required
def delete_task(task_id):
    """
    Usuwanie zadania (tylko dla administratora).
    Usuwa zadanie wraz z powiązanymi krokami i plikami.
    """
    db = get_db()
    
    # Sprawdź czy zadanie istnieje
    task = db.execute('SELECT * FROM tasks WHERE id = ?', (task_id,)).fetchone()
    if not task:
        flash('Zadanie nie istnieje')
        return redirect(url_for('index'))
    
    # Usuń powiązane pliki
    db.execute('DELETE FROM files WHERE task_id = ?', (task_id,))
    
    # Usuń powiązane kroki procesu
    db.execute('DELETE FROM process_steps WHERE task_id = ?', (task_id,))
    
    # Usuń zadanie
    db.execute('DELETE FROM tasks WHERE id = ?', (task_id,))
    db.commit()
    
    flash(f'Zadanie "{task["project_name"]}" zostało usunięte')
    return redirect(url_for('index'))


@app.route('/admin/archive')
@admin_required
def archive():
    """
    Wyświetlanie zarchiwizowanych zadań (tylko dla administratora).
    """
    db = get_db()
    
    # Pobierz wszystkie zarchiwizowane zadania
    tasks = db.execute('''
        SELECT t.*, u.username as admin_name,
               (SELECT COUNT(*) FROM process_steps WHERE task_id = t.id) as total_steps,
               (SELECT COUNT(*) FROM process_steps WHERE task_id = t.id AND status = 'zakończone') as completed_steps
        FROM tasks t 
        LEFT JOIN users u ON t.admin_id = u.id 
        WHERE t.archived = 1
        ORDER BY t.created_at DESC
    ''').fetchall()
    
    return render_template('archive.html', tasks=tasks)


@app.route('/workload')
@login_required
def workload():
    """
    Wykres obciążenia zespołu - pokazuje liczbę zadań "w toku", "nowych" i "zakończonych"
    dla każdego użytkownika lub roli.
    """
    db = get_db()
    user_role = session.get('role')
    user_id = session.get('user_id')

    # Parametry filtrowania
    view_type = request.args.get('view_type', 'user')  # 'user' lub 'role'
    status_filter = request.form.get('status_filter', request.args.get('status_filter', 'all'))
    date_from = request.form.get('date_from', request.args.get('date_from', ''))
    date_to = request.form.get('date_to', request.args.get('date_to', ''))
    filter_user_id = request.form.get('filter_user_id', request.args.get('filter_user_id', ''))

    # Pobierz wszystkich użytkowników (dla filtrowania)
    users = []
    if user_role == 'admin':
        users = db.execute('SELECT * FROM users ORDER BY username').fetchall()

    # Budowanie zapytania SQL - pobierz zadania i ich status
    query = '''
        SELECT t.id, t.status, t.admin_id, t.archived,
               u.username as assigned_username,
               u.role as assigned_role
        FROM tasks t
        LEFT JOIN users u ON t.admin_id = u.id
        WHERE (t.archived = 0 OR t.archived IS NULL)
    '''
    params = []

    # Filtrowanie po dacie utworzenia zadania
    if date_from:
        query += ' AND DATE(t.created_at) >= ?'
        params.append(date_from)
    if date_to:
        query += ' AND DATE(t.created_at) <= ?'
        params.append(date_to)

    # Dla zwykłego użytkownika pokazuj tylko zadania jego roli
    if user_role != 'admin':
        query += ''' AND t.id IN (
            SELECT DISTINCT task_id FROM process_steps 
            WHERE (',' || role_required || ',' LIKE '%,' || ? || ',%')
        )'''
        params.append(user_role)

    # Dla filtrowania po konkretnym użytkowniku (dla admina)
    if filter_user_id and user_role == 'admin':
        query += ' AND t.admin_id = ?'
        params.append(filter_user_id)

    tasks = db.execute(query, params).fetchall()

    # Grupowanie zadań
    workload_data = {}
    
    # Funkcja pomocnicza do dodawania/zliczania
    def add_to_group(group_key, group_name):
        if group_key not in workload_data:
            workload_data[group_key] = {
                'name': group_name,
                'task_count': 0,
                'in_progress': 0,
                'new': 0,
                'completed': 0
            }
        workload_data[group_key]['task_count'] += 1
        return group_key

    # Przetwarzanie zadań
    for task in tasks:
        if status_filter != 'all' and task['status'] != status_filter:
            continue
            
        if view_type == 'user':
            # Grupowanie po użytkowniku (twórca zadania / administrator)
            if task['admin_id']:
                group_key = f"user_{task['admin_id']}"
                group_name = task['assigned_username'] or f"Użytkownik {task['admin_id']}"
            else:
                group_key = "user_unknown"
                group_name = "Nieznany"
        else:
            # Grupowanie po roli (twórcy zadania)
            if task['admin_id']:
                user_info = db.execute('SELECT role FROM users WHERE id = ?', 
                                     (task['admin_id'],)).fetchone()
                if user_info:
                    group_key = f"role_{user_info['role']}"
                    group_name = user_info['role']
                else:
                    group_key = "role_unknown"
                    group_name = "Nieznana"
            else:
                group_key = "role_unknown"
                group_name = "Nieznana"

        add_to_group(group_key, group_name)
        
        if task['status'] == 'w-toku':
            workload_data[group_key]['in_progress'] += 1
        elif task['status'] == 'nowe':
            workload_data[group_key]['new'] += 1
        elif task['status'] == 'zakończone':
            workload_data[group_key]['completed'] += 1

    # Konwersja na listę i sortowanie
    workload_list = []
    for group_key, data in sorted(workload_data.items(), key=lambda x: x[1]['task_count'], reverse=True):
        # Oblicz całkowity czas dla tej grupy
        group_time_query = '''
            SELECT SUM(ps.time_actual) as total_time
            FROM process_steps ps
            JOIN tasks t ON ps.task_id = t.id
            WHERE ps.status = 'zakończone' 
            AND ps.time_actual IS NOT NULL
        '''
        group_time_params = []
        
        if view_type == 'user':
            group_time_query += ' AND t.admin_id = ?'
            if group_key == 'user_unknown':
                group_time_query += ' AND (t.admin_id IS NULL OR t.admin_id = 0)'
                group_time_params.append(0)
            else:
                try:
                    uid = int(group_key.split('_')[1])
                    group_time_params.append(uid)
                except:
                    continue
        else:
            group_time_query += ''' AND t.id IN (
                SELECT task_id FROM process_steps WHERE task_id IN (
                    SELECT id FROM tasks WHERE admin_id IN (
                        SELECT id FROM users WHERE role = ?
                    )
                )
            )'''
            group_time_params.append(data['name'])
        
        if date_from:
            group_time_query += ' AND DATE(ps.completed_at) >= ?'
            group_time_params.append(date_from)
        if date_to:
            group_time_query += ' AND DATE(ps.completed_at) <= ?'
            group_time_params.append(date_to)
        
        time_result = db.execute(group_time_query, group_time_params).fetchone()
        group_total_hours = float(time_result['total_time']) if time_result and time_result['total_time'] else 0
        
        data['total_hours'] = group_total_hours
        workload_list.append(data)

    # Obliczanie statystyk
    total_tasks = sum(item['task_count'] for item in workload_list)
    active_tasks = sum(item['in_progress'] for item in workload_list)
    completed_tasks = sum(item['completed'] for item in workload_list)

    # Obliczanie całkowitego czasu wykonania (z zakończonych kroków)
    time_query = '''
        SELECT SUM(time_actual) as total_time
        FROM process_steps
        WHERE status = 'zakończone' AND time_actual IS NOT NULL
    '''
    time_params = []
    
    if date_from:
        time_query += ' AND DATE(completed_at) >= ?'
        time_params.append(date_from)
    if date_to:
        time_query += ' AND DATE(completed_at) <= ?'
        time_params.append(date_to)

    time_result = db.execute(time_query, time_params).fetchone()
    total_hours = float(time_result['total_time']) if time_result and time_result['total_time'] else 0

    workload_stats = {
        'total_tasks': total_tasks,
        'active_tasks': active_tasks,
        'completed_tasks': completed_tasks,
        'total_hours': total_hours
    }

    return render_template('workload.html',
                         workload_data=workload_list,
                         workload_stats=workload_stats,
                         view_type=view_type,
                         status_filter=status_filter,
                         date_from=date_from,
                         date_to=date_to,
                         filter_user_id=filter_user_id,
                         users=users)


@app.route('/workload/export/pdf', methods=['POST'])
@login_required
def export_workload_pdf():
    """
    Eksport wykresu obciążenia zespołu do formatu PDF.
    """
    db = get_db()
    user_role = session.get('role')
    user_id = session.get('user_id')

    view_type = request.form.get('view_type', 'user')
    status_filter = request.form.get('status_filter', 'all')
    date_from = request.form.get('date_from', '')
    date_to = request.form.get('date_to', '')
    filter_user_id = request.form.get('filter_user_id', '')

    # Pobierz dane dla PDF (podobnie jak w workload)
    query = '''
        SELECT t.id, t.status, t.admin_id, t.archived,
               u.username as assigned_username,
               u.role as assigned_role
        FROM tasks t
        LEFT JOIN users u ON t.admin_id = u.id
        WHERE (t.archived = 0 OR t.archived IS NULL)
    '''
    params = []

    if date_from:
        query += ' AND DATE(t.created_at) >= ?'
        params.append(date_from)
    if date_to:
        query += ' AND DATE(t.created_at) <= ?'
        params.append(date_to)

    if user_role != 'admin':
        query += ''' AND t.id IN (
            SELECT DISTINCT task_id FROM process_steps 
            WHERE (',' || role_required || ',' LIKE '%,' || ? || ',%')
        )'''
        params.append(user_role)

    if filter_user_id and user_role == 'admin':
        query += ' AND t.admin_id = ?'
        params.append(filter_user_id)

    tasks = db.execute(query, params).fetchall()

    # Grupowanie danych
    workload_data = {}
    for task in tasks:
        if status_filter != 'all' and task['status'] != status_filter:
            continue

        if view_type == 'user':
            if task['admin_id']:
                group_key = f"user_{task['admin_id']}"
                group_name = task['assigned_username'] or f"Użytkownik {task['admin_id']}"
            else:
                group_key = "user_unknown"
                group_name = "Nieznany"
        else:
            if task['admin_id']:
                user_info = db.execute('SELECT role FROM users WHERE id = ?', 
                                     (task['admin_id'],)).fetchone()
                if user_info:
                    group_key = f"role_{user_info['role']}"
                    group_name = user_info['role']
                else:
                    group_key = "role_unknown"
                    group_name = "Nieznana"
            else:
                group_key = "role_unknown"
                group_name = "Nieznana"

        if group_key not in workload_data:
            workload_data[group_key] = {
                'name': group_name,
                'total': 0,
                'in_progress': 0,
                'new': 0,
                'completed': 0
            }
        workload_data[group_key]['total'] += 1
        
        if task['status'] == 'w-toku':
            workload_data[group_key]['in_progress'] += 1
        elif task['status'] == 'nowe':
            workload_data[group_key]['new'] += 1
        elif task['status'] == 'zakończone':
            workload_data[group_key]['completed'] += 1

    workload_list = sorted(list(workload_data.values()), key=lambda x: x['total'], reverse=True)

    # Tworzenie PDF
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", size=10)

    # Nagłówek
    pdf.set_font("Helvetica", 'B', 16)
    pdf.cell(0, 10, 'Wykres obciazenia zespolu', 0, 1, 'C')
    pdf.set_font("Helvetica", size=10)
    pdf.cell(0, 10, 'Wygenerowano: {}'.format(datetime.now().strftime("%Y-%m-%d %H:%M:%S")), 0, 1, 'C')
    pdf.ln(5)

    # Filtry
    view_label = "Uzytkownicy" if view_type == 'user' else "Role"
    pdf.cell(0, 10, 'Podzial: {}'.format(view_label), 0, 1)
    status_label = "Wszystkie" if status_filter == 'all' else status_filter
    pdf.cell(0, 10, 'Status: {}'.format(status_label), 0, 1)
    if date_from or date_to:
        pdf.cell(0, 10, 'Data: {} - {}'.format(date_from or "...", date_to or "..."), 0, 1)
    pdf.ln(10)

    # Tabela
    if workload_list:
        col_widths = [80, 20, 20, 25, 25]
        headers = ['Uzytkownik/Rola', 'Lacznie', 'W toku', 'Nowe', 'Zakonczone']

        pdf.set_fill_color(200, 220, 255)
        for i, header in enumerate(headers):
            pdf.cell(col_widths[i], 10, header, 1, 0, 'C', 1)
        pdf.ln()

        pdf.set_fill_color(255, 255, 255)
        for item in workload_list:
            name = item['name'].encode('latin-1', 'ignore').decode('latin-1')[:25] if item['name'] else ''
            pdf.cell(col_widths[0], 10, name, 1, 0)
            pdf.cell(col_widths[1], 10, str(item['total']), 1, 0, 'C')
            pdf.cell(col_widths[2], 10, str(item['in_progress']), 1, 0, 'C')
            pdf.cell(col_widths[3], 10, str(item['new']), 1, 0, 'C')
            pdf.cell(col_widths[4], 10, str(item['completed']), 1, 1, 'C')

        # Podsumowanie
        total_all = sum(item['total'] for item in workload_list)
        total_progress = sum(item['in_progress'] for item in workload_list)
        pdf.ln(10)
        pdf.cell(0, 10, 'Lacznie zadan: {}, W toku: {}'.format(total_all, total_progress), 0, 1)
    else:
        pdf.cell(0, 10, 'Brak danych do wyswietlenia', 0, 1)

    pdf_output = pdf.output(dest='S').encode('latin-1')
    buffer = io.BytesIO(pdf_output)
    buffer.seek(0)

    return send_file(buffer, as_attachment=True,
                   download_name='obciazenie_zespolu_{}.pdf'.format(datetime.now().strftime("%Y%m%d_%H%M%S")),
                   mimetype='application/pdf')


@app.route('/workload/export/excel', methods=['POST'])
@login_required
def export_workload_excel():
    """
    Eksport wykresu obciążenia zespołu do formatu Excel.
    """
    db = get_db()
    user_role = session.get('role')
    user_id = session.get('user_id')

    view_type = request.form.get('view_type', 'user')
    status_filter = request.form.get('status_filter', 'all')
    date_from = request.form.get('date_from', '')
    date_to = request.form.get('date_to', '')
    filter_user_id = request.form.get('filter_user_id', '')

    # Pobierz dane (podobnie jak w PDF)
    query = '''
        SELECT t.id, t.status, t.admin_id, t.archived,
               u.username as assigned_username,
               u.role as assigned_role
        FROM tasks t
        LEFT JOIN users u ON t.admin_id = u.id
        WHERE (t.archived = 0 OR t.archived IS NULL)
    '''
    params = []

    if date_from:
        query += ' AND DATE(t.created_at) >= ?'
        params.append(date_from)
    if date_to:
        query += ' AND DATE(t.created_at) <= ?'
        params.append(date_to)

    if user_role != 'admin':
        query += ''' AND t.id IN (
            SELECT DISTINCT task_id FROM process_steps 
            WHERE (',' || role_required || ',' LIKE '%,' || ? || ',%')
        )'''
        params.append(user_role)

    if filter_user_id and user_role == 'admin':
        query += ' AND t.admin_id = ?'
        params.append(filter_user_id)

    tasks = db.execute(query, params).fetchall()

    # Grupowanie
    workload_data = {}
    for task in tasks:
        if status_filter != 'all' and task['status'] != status_filter:
            continue

        if view_type == 'user':
            if task['admin_id']:
                group_key = f"user_{task['admin_id']}"
                group_name = task['assigned_username'] or f"Użytkownik {task['admin_id']}"
            else:
                group_key = "user_unknown"
                group_name = "Nieznany"
        else:
            if task['admin_id']:
                user_info = db.execute('SELECT role FROM users WHERE id = ?', 
                                     (task['admin_id'],)).fetchone()
                if user_info:
                    group_key = f"role_{user_info['role']}"
                    group_name = user_info['role']
                else:
                    group_key = "role_unknown"
                    group_name = "Nieznana"
            else:
                group_key = "role_unknown"
                group_name = "Nieznana"

        if group_key not in workload_data:
            workload_data[group_key] = {
                'name': group_name,
                'total': 0,
                'in_progress': 0,
                'new': 0,
                'completed': 0
            }
        workload_data[group_key]['total'] += 1
        
        if task['status'] == 'w-toku':
            workload_data[group_key]['in_progress'] += 1
        elif task['status'] == 'nowe':
            workload_data[group_key]['new'] += 1
        elif task['status'] == 'zakończone':
            workload_data[group_key]['completed'] += 1

    workload_list = sorted(list(workload_data.values()), key=lambda x: x['total'], reverse=True)

    # Tworzenie Excela
    wb = Workbook()
    ws = wb.active
    ws.title = "Obciazenie zespolu"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    headers = ['Uzytkownik/Rola', 'Lacznie zadan', 'W toku', 'Nowe', 'Zakonczone']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, item in enumerate(workload_list, 2):
        name = item['name']
        ws.cell(row=row_idx, column=1, value=name)
        ws.cell(row=row_idx, column=2, value=item['total'])
        ws.cell(row=row_idx, column=3, value=item['in_progress'])
        ws.cell(row=row_idx, column=4, value=item['new'])
        ws.cell(row=row_idx, column=5, value=item['completed'])

    total_row = len(workload_list) + 3
    ws.cell(row=total_row, column=1, value="Lacznie:")
    ws.cell(row=total_row, column=2, value=sum(item['total'] for item in workload_list))

    ws.column_dimensions['A'].width = 25
    for col in ['B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 15

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(buffer, as_attachment=True,
                   download_name='obciazenie_zespolu_{}.xlsx'.format(datetime.now().strftime("%Y%m%d_%H%M%S")),
                   mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    """
    Profil użytkownika - edycja danych i zarządzanie umiejętnościami.
    """
    db = get_db()
    user_id = session.get('user_id')
    
    if request.method == 'POST':
        # Aktualizacja danych podstawowych
        full_name = request.form.get('full_name', '')
        db.execute('UPDATE users SET full_name = ? WHERE id = ?', (full_name, user_id))
        
        # Aktualizacja umiejętności (tagów)
        selected_skills = request.form.getlist('skills')
        
        # Usuń istniejące powiązania
        db.execute('DELETE FROM user_skills WHERE user_id = ?', (user_id,))
        
        # Dodaj nowe powiązania
        for skill_id in selected_skills:
            try:
                db.execute('INSERT INTO user_skills (user_id, skill_id) VALUES (?, ?)',
                          (user_id, int(skill_id)))
            except sqlite3.IntegrityError:
                pass  # skill już istnieje (IGNORE)
        
        db.commit()
        flash('Profil zaktualizowany')
        return redirect(url_for('profile'))
    
    # GET - pobierz dane użytkownika i umiejętności
    user = db.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
    
    # Pobierz wszystkie dostępne umiejętności
    all_skills = db.execute('SELECT * FROM skills ORDER BY category, name').fetchall()
    
    # Pobierz umiejętności użytkownika
    user_skill_ids = [row['skill_id'] for row in 
                      db.execute('SELECT skill_id FROM user_skills WHERE user_id = ?', (user_id,)).fetchall()]
    
    return render_template('profile.html', 
                         user=user, 
                         all_skills=all_skills,
                         user_skill_ids=user_skill_ids)


@app.route('/api/suggested-performers/<int:step_id>')
@login_required
def suggested_performers(step_id):
    """
    API - zwraca listę użytkowników sugerowanych do wykonania kroku
    na podstawie ról i umiejętności.
    """
    db = get_db()
    
    # Pobierz krok
    step = db.execute('SELECT * FROM process_steps WHERE id = ?', (step_id,)).fetchone()
    if not step:
        return jsonify({'error': 'Step not found'}), 404
    
    # Znajdź użytkowników z odpowiednią rolą
    role_param = step['role_required']
    # Jeśli rola to lista oddzielona przecinkami, bierzemy pierwszą
    if ',' in role_param:
        role_param = role_param.split(',')[0].strip()
    
    potential_users = db.execute(
        "SELECT * FROM users WHERE role = ? OR role = 'admin'", 
        (role_param,)
    ).fetchall()
    
    # Jeśli krok ma wymaganą umiejętność, dodaj filtr
    required_skill = step['required_skill']
    suggested = []
    
    for user in potential_users:
        # Sprawdź czy użytkownik ma wymaganą umiejętność
        if required_skill:
            skill_check = db.execute('''
                SELECT us.id FROM user_skills us
                JOIN skills s ON us.skill_id = s.id
                WHERE us.user_id = ? AND s.name = ?
            ''', (user['id'], required_skill)).fetchone()
            if not skill_check:
                continue  # Brak wymaganej umiejętności
        
        # Policz ile zadań ma użytkownik w toku
        active_tasks = db.execute('''
            SELECT COUNT(DISTINCT t.id) as count
            FROM tasks t
            JOIN process_steps ps ON t.id = ps.task_id
            WHERE t.status = 'w-toku' 
            AND (',' || ps.role_required || ',' LIKE '%,' || ? || ',%')
        ''', (user['role'],)).fetchone()['count']
        
        suggested.append({
            'id': user['id'],
            'username': user['username'],
            'full_name': user['full_name'] or user['username'],
            'role': user['role'],
            'active_tasks': active_tasks,
            'has_skill': required_skill is None or True
        })
    
    # Sortuj po liczbie aktywnych zadań (najmniej zajęty na górze)
    suggested.sort(key=lambda x: x['active_tasks'])
    
    return jsonify(suggested)


@app.route('/admin/skills', methods=['GET', 'POST'])
@admin_required
def manage_skills():
    """
    Zarządzanie umiejętnościami (matryca kompetencji) - tylko dla administratora.
    GET: Wyświetla listę umiejętności i użytkowników z przypisanymi kompetencjami
    POST: Obsługuje dodawanie/edycję/uszuwanie umiejętności i przypisywanie do użytkowników
    """
    db = get_db()
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        # Dodanie nowej umiejętności
        if action == 'add_skill':
            name = request.form.get('skill_name', '').strip()
            description = request.form.get('skill_description', '').strip()
            category = request.form.get('skill_category', '').strip()
            
            if name:
                try:
                    db.execute('INSERT INTO skills (name, description, category) VALUES (?, ?, ?)',
                              (name, description, category))
                    db.commit()
                    flash(f'Umiejętność "{name}" dodana pomyślnie')
                except sqlite3.IntegrityError:
                    flash(f'Umiejętność "{name}" już istnieje')
            else:
                flash('Nazwa umiejętności jest wymagana')
        
        # Przypisanie umiejętności do użytkownika
        elif action == 'assign_skill':
            user_id = request.form.get('user_id')
            skill_id = request.form.get('skill_id')
            proficiency = request.form.get('proficiency_level', 3)
            
            if user_id and skill_id:
                try:
                    db.execute('INSERT OR REPLACE INTO user_skills (user_id, skill_id, proficiency_level) VALUES (?, ?, ?)',
                              (user_id, skill_id, proficiency))
                    db.commit()
                    flash('Umiejętność przypisana pomyślnie')
                except Exception as e:
                    flash(f'Błąd: {str(e)}')
            else:
                flash('Wybierz użytkownika i umiejętność')
        
        # Usunięcie umiejętności z użytkownika
        elif action == 'remove_skill':
            user_id = request.form.get('user_id')
            skill_id = request.form.get('skill_id')
            
            if user_id and skill_id:
                db.execute('DELETE FROM user_skills WHERE user_id = ? AND skill_id = ?', (user_id, skill_id))
                db.commit()
                flash('Umiejętność usunięta pomyślnie')
        
        # Usunięcie umiejętności ze słownika
        elif action == 'delete_skill':
            skill_id = request.form.get('skill_id')
            if skill_id:
                # Usuń powiązania najpierw (ON DELETE CASCADE powinno zadziałać)
                db.execute('DELETE FROM user_skills WHERE skill_id = ?', (skill_id,))
                db.execute('DELETE FROM skills WHERE id = ?', (skill_id,))
                db.commit()
                flash('Umiejętność usunięta ze słownika')
    
    # GET - pobierz dane do wyświetlenia
    skills = db.execute('SELECT * FROM skills ORDER BY category, name').fetchall()
    users = db.execute('SELECT * FROM users ORDER BY username').fetchall()
    
    # Pobierz wszystkie przypisania user_skills z nazwami
    assignments = db.execute('''
        SELECT us.*, s.name as skill_name, s.category, u.username as user_name
        FROM user_skills us
        JOIN skills s ON us.skill_id = s.id
        JOIN users u ON us.user_id = u.id
        ORDER BY u.username, s.name
    ''').fetchall()
    
    return render_template('manage_skills.html', 
                         skills=skills, 
                         users=users,
                         assignments=assignments)


@app.route('/admin/step/<int:step_id>/set-skill', methods=['POST'])
@admin_required
def set_step_skill(step_id):
    """
    Ustawienie wymaganej umiejętności dla kroku (admin).
    """
    db = get_db()
    skill_id = request.form.get('skill_id', '')
    
    # Pobierz nazwę umiejętności jeśli skill_id nie jest pusty
    if skill_id:
        skill = db.execute('SELECT name FROM skills WHERE id = ?', (skill_id,)).fetchone()
        if skill:
            db.execute('UPDATE process_steps SET required_skill = ? WHERE id = ?',
                      (skill['name'], step_id))
            flash(f'Ustawiono wymaganą umiejętność: {skill["name"]}')
    else:
        # Usuń wymaganie umiejętności
        db.execute('UPDATE process_steps SET required_skill = NULL WHERE id = ?', (step_id,))
        flash('Usunięto wymaganą umiejętność')
    
    db.commit()
    return redirect(url_for('manage_steps'))


if __name__ == '__main__':
    """
    Główny punkt wejścia aplikacji.
    Inicjalizuje bazę danych i uruchamia serwer Flask.
    - host='0.0.0.0' - serwer dostępny z innych komputerów w sieci LAN
    - port=5000 - port na którym działa aplikacja
    - debug=True - tryb debugowania (auto-reładowanie przy zmianach kodu)
    """
    init_db()
    app.run(host='0.0.0.0', port=5000, debug=True)
