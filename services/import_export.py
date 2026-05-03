import io
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from flask import current_app
from datetime import datetime
from app.models import Project, Step, User, Skill, Role
from app import db
from sqlalchemy.exc import SQLAlchemyError


def export_projects_to_excel(projects_query=None):
    """Eksport projektów do formatu Excel (XLSX).
    
    Args:
        projects_query: Zapytanie SQLAlchemy z projektami (jeśli None, eksportuje wszystkie)
        
    Returns:
        BytesIO: Plik Excel w pamięci
    """
    if projects_query is None:
        projects = Project.query.all()
    else:
        projects = projects_query.all()
    
    data = []
    for p in projects:
        data.append({
            'id': p.id,
            'nazwa': p.name,
            'opis': p.description or '',
            'status': p.status,
            'priorytet': p.priority,
            'prioriety_etykieta': p.priority_label,
            'data_utworzenia': p.created_at.strftime('%Y-%m-%d %H:%M:%S') if p.created_at else '',
            'termin_wykonania': p.due_date.strftime('%Y-%m-%d') if p.due_date else '',
            'procent_ukonczenia': p.completion_percentage or 0,
            'liczba_zadan': p.steps.count(),
            'utworzony_przez': p.creator.full_name if p.creator else '',
            'szablon_projektu': p.project_template.name if p.project_template else ''
        })
    
    df = pd.DataFrame(data)
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Projekty', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Projekty']
        
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        column_widths = {'A': 5, 'B': 30, 'C': 40, 'D': 12, 'E': 10, 'F': 12,
                        'G': 20, 'H': 15, 'I': 18, 'J': 15, 'K': 20, 'L': 25}
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, max_col=worksheet.max_column):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
    
    output.seek(0)
    return output


def export_steps_to_excel(steps_query=None):
    """Eksport zadań (kroków) do formatu Excel (XLSX).
    
    Args:
        steps_query: Zapytanie SQLAlchemy z zadaniami (jeśli None, eksportuje wszystkie)
        
    Returns:
        BytesIO: Plik Excel w pamięci
    """
    if steps_query is None:
        steps = Step.query.all()
    else:
        steps = steps_query.all()
    
    data = []
    for s in steps:
        data.append({
            'id': s.id,
            'nazwa': s.name,
            'projekt': s.project.name if s.project else '',
            'typ_kroku': s.step_type or '',
            'przypisany_do': s.assignee.full_name if s.assignee else '',
            'data_utworzenia': s.created_at.strftime('%Y-%m-%d %H:%M:%S') if s.created_at else '',
            'termin_wykonania': s.due_date.strftime('%Y-%m-%d') if s.due_date else '',
            'priorytet': s.priority,
            'status': s.status,
            'szacowany_czas_min': s.estimated_time or 0,
            'rzeczywisty_czas_min': s.actual_time or 0,
            'opis': s.description or '',
            'notatki': s.notes or '',
            'utworzony_przez': s.creator.full_name if s.creator else '',
            'przeterminowany': 'TAK' if s.is_overdue else 'NIE'
        })
    
    df = pd.DataFrame(data)
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Zadania', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Zadania']
        
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        column_widths = {'A': 5, 'B': 30, 'C': 20, 'D': 20, 'E': 25, 'F': 20,
                        'G': 15, 'H': 12, 'I': 12, 'J': 15, 'K': 15,
                        'L': 40, 'M': 30, 'N': 20, 'O': 12}
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, max_col=worksheet.max_column):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
    
    output.seek(0)
    return output


def export_users_to_excel():
    """Eksport użytkowników z ich umiejętnościami do formatu Excel (XLSX).
    
    Returns:
        BytesIO: Plik Excel w pamięci
    """
    users = User.query.all()
    
    data = []
    for u in users:
        skills_list = ', '.join([s.name for s in u.skills]) if u.skills else 'Brak'
        roles_list = ', '.join([r.display_name for r in u.roles]) if u.roles else 'Zespół'
        data.append({
            'id': u.id,
            'nazwa_uzytkownika': u.username,
            'imie': u.first_name or '',
            'nazwisko': u.last_name or '',
            'email': u.email,
            'role': roles_list,  # Changed: show all roles
            'status': 'Aktywny' if u.is_active else 'Nieaktywny',
            'umiejetnosci': skills_list,
            'liczba_zadan_przypisanych': Step.query.filter_by(assigned_to=u.id, status='in_progress').count(),
            'liczba_zadan_ukonczonych': Step.query.filter_by(assigned_to=u.id, status='completed').count(),
            'data_utworzenia': u.created_at.strftime('%Y-%m-%d %H:%M:%S') if u.created_at else ''
        })
    
    df = pd.DataFrame(data)
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Użytkownicy', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Użytkownicy']
        
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        column_widths = {'A': 5, 'B': 15, 'C': 15, 'D': 15, 'E': 25,
                        'F': 15, 'G': 10, 'H': 40, 'I': 20, 'J': 20, 'K': 20}
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, max_col=worksheet.max_column):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
    
    output.seek(0)
    return output


def _get_priority_number(priority_str):
    priority_map = {'low': 1, 'medium': 2, 'high': 3, 'urgent': 4}
    if isinstance(priority_str, str):
        return priority_map.get(priority_str.lower(), 2)
    try:
        val = int(priority_str)
        return max(1, min(4, val))
    except (ValueError, TypeError):
        return 2


def _parse_date(date_str):
    if not date_str:
        return None
    if isinstance(date_str, datetime):
        return date_str
    date_str = str(date_str).strip()
    formats = ['%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y', '%d/%m/%Y', '%Y/%m/%d']
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    return None


def import_projects_from_excel(file_stream, user_id):
    """Importuje projekty z pliku Excel.
    
    Tworzy nowe projekty lub aktualizuje istniejące (na podstawie nazwy).
    
    Args:
        file_stream: Strumień pliku Excel
        user_id: ID użytkownika importującego
        
    Returns:
        dict: Wynik importu z liczbami i błędami
    """
    try:
        df = pd.read_excel(file_stream, engine='openpyxl')
    except Exception as e:
        return {'success': False, 'imported': 0, 'updated': 0,
                'errors': [f'Błąd odczytu pliku: {str(e)}']}
    
    column_mapping = {'nazwa': 'name', 'opis': 'description', 'status': 'status',
                     'priorytet': 'priority', 'termin_wykonania': 'due_date',
                     'nazwa_projektu': 'name', 'projekt': 'project'}
    df.rename(columns=column_mapping, inplace=True)
    
    required_fields = ['name']
    missing_fields = [f for f in required_fields if f not in df.columns]
    if missing_fields:
        return {'success': False, 'imported': 0, 'updated': 0,
                'errors': [f'Brak wymaganych kolumn: {", ".join(missing_fields)}']}
    
    results = {'success': True, 'imported': 0, 'updated': 0, 'errors': []}
    valid_statuses = ['active', 'completed', 'archived']
    
    for idx, row in df.iterrows():
        row_num = idx + 2
        try:
            name = str(row.get('name', '')).strip()
            if not name:
                results['errors'].append(f'Wiersz {row_num}: Brak nazwy - pominięto')
                continue
            
            existing = Project.query.filter_by(name=name).first()
            description = str(row.get('description', '')).strip()
            priority = row.get('priority', 2)
            status = str(row.get('status', 'active')).lower().strip()
            due_date = _parse_date(row.get('due_date'))
            
            if status not in valid_statuses:
                status = 'active'
            
            if existing:
                existing.description = description
                existing.priority = _get_priority_number(priority)
                existing.status = status
                existing.due_date = due_date
                results['updated'] += 1
            else:
                project = Project(name=name, description=description,
                                priority=_get_priority_number(priority),
                                status=status, due_date=due_date,
                                created_by=user_id)
                db.session.add(project)
                results['imported'] += 1
            
            db.session.flush()
        except Exception as e:
            db.session.rollback()
            results['errors'].append(f'Wiersz {row_num}: Błąd: {str(e)}')
    
    if results['imported'] > 0 or results['updated'] > 0:
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            results['success'] = False
            results['errors'].append(f'Błąd zapisu: {str(e)}')
    
    return results


def import_steps_from_excel(file_stream, user_id):
    """Importuje zadania (kroki) z pliku Excel.
    
    Args:
        file_stream: Strumień pliku Excel
        user_id: ID użytkownika importującego
        
    Returns:
        dict: Wynik importu z liczbami i błędami
    """
    try:
        df = pd.read_excel(file_stream, engine='openpyxl')
    except Exception as e:
        return {'success': False, 'imported': 0, 'updated': 0,
                'errors': [f'Błąd odczytu pliku: {str(e)}']}
    
    column_mapping = {'nazwa': 'name', 'projekt': 'project', 'typ': 'step_type',
                     'typ_kroku': 'step_type', 'przypisany': 'assigned_to',
                     'termin': 'due_date', 'priorytet': 'priority', 'status': 'status',
                     'szacowany_czas': 'estimated_time', 'opis': 'description',
                     'notatki': 'notes'}
    df.rename(columns=column_mapping, inplace=True)
    
    required_fields = ['name']
    missing_fields = [f for f in required_fields if f not in df.columns]
    if missing_fields:
        return {'success': False, 'imported': 0, 'updated': 0,
                'errors': [f'Brak wymaganych kolumn: {", ".join(missing_fields)}']}
    
    status_mapping = {'do_zrobienia': 'pending', 'pending': 'pending',
                     'w_trakcie': 'in_progress', 'in_progress': 'in_progress',
                     'zamkniete': 'completed', 'zakończone': 'completed',
                     'completed': 'completed'}
    
    all_users = {u.username.lower(): u.id for u in User.query.all()}
    all_users.update({u.email.lower(): u.id for u in User.query.all()})
    all_users.update({str(u.id): u.id for u in User.query.all()})
    
    all_projects = {p.name.lower(): p.id for p in Project.query.all()}
    all_projects.update({str(p.id): p.id for p in Project.query.all()})
    
    results = {'success': True, 'imported': 0, 'updated': 0, 'errors': []}
    priority_mapping = {'low': 'normal', 'medium': 'normal', 'high': 'high',
                       'urgent': 'urgent', 'very_urgent': 'very_urgent',
                       'very urgent': 'very_urgent'}
    
    for idx, row in df.iterrows():
        row_num = idx + 2
        try:
            name = str(row.get('name', '')).strip()
            if not name:
                results['errors'].append(f'Wiersz {row_num}: Brak nazwy - pominięto')
                continue
            
            project_id = None
            project_ref = str(row.get('project', '')).strip()
            if project_ref:
                project_id = all_projects.get(project_ref.lower())
                if not project_id:
                    try:
                        pid = int(float(project_ref))
                        project_id = all_projects.get(str(pid))
                    except ValueError:
                        pass
                    if not project_id:
                        results['errors'].append(f'Wiersz {row_num}: Projekt "{project_ref}" nie istnieje')
            
            assigned_to = None
            assigned_ref = str(row.get('assigned_to', '')).strip()
            if assigned_ref:
                assigned_to = all_users.get(assigned_ref.lower())
                if not assigned_to:
                    try:
                        uid = int(float(assigned_ref))
                        assigned_to = all_users.get(str(uid))
                    except ValueError:
                        pass
                    if not assigned_to:
                        results['errors'].append(f'Wiersz {row_num}: Użytkownik "{assigned_ref}" nie istnieje')
            
            step_type = str(row.get('step_type', '')).strip()
            priority_raw = str(row.get('priority', 'normal')).strip().lower()
            priority = priority_mapping.get(priority_raw, 'normal')
            status_raw = str(row.get('status', 'pending')).strip().lower()
            status = status_mapping.get(status_raw, 'pending')
            
            estimated_time = None
            try:
                est = row.get('estimated_time')
                if est is not None and str(est).strip():
                    estimated_time = int(float(str(est).strip()))
            except (ValueError, TypeError):
                pass
            
            due_date = _parse_date(row.get('due_date'))
            description = str(row.get('description', '')).strip()
            notes = str(row.get('notes', '')).strip()
            
            query = Step.query.filter_by(name=name)
            if project_id:
                query = query.filter_by(project_id=project_id)
            existing = query.first()
            
            if existing:
                existing.description = description
                existing.step_type = step_type
                existing.assigned_to = assigned_to
                existing.estimated_time = estimated_time
                existing.due_date = due_date
                existing.priority = priority
                existing.status = status
                existing.notes = notes
                results['updated'] += 1
            else:
                step = Step(name=name, description=description,
                          step_type=step_type, assigned_to=assigned_to,
                          created_by=user_id, estimated_time=estimated_time,
                          due_date=due_date, priority=priority, status=status,
                          notes=notes, project_id=project_id)
                db.session.add(step)
                results['imported'] += 1
            
            db.session.flush()
        except Exception as e:
            db.session.rollback()
            results['errors'].append(f'Wiersz {row_num}: Błąd: {str(e)}')
    
    if results['imported'] > 0 or results['updated'] > 0:
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            results['success'] = False
            results['errors'].append(f'Błąd zapisu: {str(e)}')
    
    return results


def import_users_from_excel(file_stream, user_id):
    """Importuje użytkowników z pliku Excel.
    
    Args:
        file_stream: Strumień pliku Excel
        user_id: ID użytkownika importującego
        
    Returns:
        dict: Wynik importu z liczbami i błędami
    """
    try:
        df = pd.read_excel(file_stream, engine='openpyxl')
    except Exception as e:
        return {'success': False, 'imported': 0, 'updated': 0,
                'errors': [f'Błąd odczytu pliku: {str(e)}']}
    
    column_mapping = {'nazwa_uzytkownika': 'username', 'username': 'username',
                     'imie': 'first_name', 'nazwisko': 'last_name',
                     'email': 'email', 'rola': 'role',
                     'password': 'password', 'haslo': 'password',
                     'skille': 'skills', 'umiejetnosci': 'skills'}
    df.rename(columns=column_mapping, inplace=True)
    
    required_fields = ['username', 'email']
    missing_fields = [f for f in required_fields if f not in df.columns]
    if missing_fields:
        return {'success': False, 'imported': 0, 'updated': 0,
                'errors': [f'Brak wymaganych kolumn: {", ".join(missing_fields)}']}
    
    existing_users = {u.username.lower(): u for u in User.query.all()}
    existing_users.update({u.email.lower(): u for u in User.query.all()})
    all_skills = {s.name.lower(): s for s in Skill.query.all()}
    all_roles = {r.name.lower(): r for r in Role.query.all()}
    
    results = {'success': True, 'imported': 0, 'updated': 0, 'errors': []}
    
    from werkzeug.security import generate_password_hash
    
    for idx, row in df.iterrows():
        row_num = idx + 2
        try:
            username = str(row.get('username', '')).strip()
            email = str(row.get('email', '')).strip()
            
            if not username or not email:
                results['errors'].append(f'Wiersz {row_num}: Brak username/email - pominięto')
                continue
            
            existing = existing_users.get(username.lower()) or existing_users.get(email.lower())
            first_name = str(row.get('first_name', '')).strip()
            last_name = str(row.get('last_name', '')).strip()
            
            # Parse roles - can be comma-separated list of role names
            roles_raw = str(row.get('role', '')).strip()
            # For backwards compatibility, also check 'roles' column
            if not roles_raw:
                roles_raw = str(row.get('roles', '')).strip()
            
            user_roles = []
            if roles_raw:
                # Split by comma, semicolon, or newline
                for role_name in re.split(r'[;,\n]', roles_raw):
                    role_name = role_name.strip().lower()
                    if role_name and role_name in all_roles:
                        user_roles.append(all_roles[role_name])
            # If no roles specified, assign default 'zespół' role
            if not user_roles:
                default_role = all_roles.get('zespół')
                if default_role:
                    user_roles.append(default_role)
            
            password_raw = str(row.get('password', '')).strip()
            if not password_raw:
                password_raw = 'changeme123'
            
            skills_raw = str(row.get('skills', '')).strip()
            skills_list = []
            if skills_raw:
                for skill_name in re.split(r'[;,\n]', skills_raw):
                    skill_name = skill_name.strip()
                    if skill_name:
                        skill = all_skills.get(skill_name.lower())
                        if skill:
                            skills_list.append(skill)
            
            if existing:
                existing.first_name = first_name or existing.first_name
                existing.last_name = last_name or existing.last_name
                existing.email = email
                # Update roles
                existing.roles = user_roles
                if password_raw and password_raw != 'changeme123':
                    existing.password_hash = generate_password_hash(password_raw)
                existing.skills = skills_list
                results['updated'] += 1
            else:
                new_user = User(username=username, email=email,
                              first_name=first_name, last_name=last_name,
                              password_hash=generate_password_hash(password_raw))
                new_user.roles = user_roles
                new_user.skills = skills_list
                db.session.add(new_user)
                results['imported'] += 1
            
            db.session.flush()
        except Exception as e:
            db.session.rollback()
            results['errors'].append(f'Wiersz {row_num}: Błąd: {str(e)}')
    
    if results['imported'] > 0 or results['updated'] > 0:
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            results['success'] = False
            results['errors'].append(f'Błąd zapisu: {str(e)}')
    
    return results


def generate_projects_template():
    output = io.BytesIO()
    template_data = {
        'nazwa': ['Projekt Alpha', 'Projekt Beta', ''],
        'opis': ['Pierwsza feta projektu', 'Dokumentacja techniczna', ''],
        'status': ['active', 'active', 'completed'],
        'priorytet': [3, 2, 1],
        'termin_wykonania': ['2024-12-31', '2024-11-30', '']
    }
    df = pd.DataFrame(template_data)
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Szablon_Projektów', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Szablon_Projektów']
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        column_widths = {'A': 25, 'B': 35, 'C': 12, 'D': 12, 'E': 18}
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, max_col=worksheet.max_column):
            for cell in row:
                cell.border = border
    output.seek(0)
    return output


def generate_users_template():
    output = io.BytesIO()
    all_skills = Skill.query.all()
    skills_example = ', '.join([s.name for s in all_skills[:5]]) if all_skills else 'Python, SQL'
    # Get role examples
    from app.models.role import Role
    all_roles = Role.query.all()
    roles_example = 'zespół' if not all_roles else ', '.join([r.name for r in all_roles[:3]])
    template_data = {
        'username': ['jan.kowalski', 'anna.nowak', ''],
        'email': ['jan.kowalski@company.com', 'anna.nowak@company.com', ''],
        'first_name': ['Jan', 'Anna', ''],
        'last_name': ['Kowalski', 'Nowak', ''],
        'role': [roles_example, 'admin', ''],
        'password': ['haslo123', 'haslo456', ''],
        'skills': [skills_example, '', '']
    }
    df = pd.DataFrame(template_data)
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Szablon_Uzytkowników', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Szablon_Uzytkowników']
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        column_widths = {'A': 15, 'B': 25, 'C': 12, 'D': 12, 'E': 10, 'F': 12, 'G': 35}
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, max_col=worksheet.max_column):
            for cell in row:
                cell.border = border
    output.seek(0)
    return output


def generate_steps_template():
    output = io.BytesIO()
    template_data = {
        'nazwa': ['Implementacja modelu 3D', 'Review dokumentacji', 'Testy akceptacyjne'],
        'projekt': ['Projekt Alpha', 'Projekt Beta', ''],
        'typ_kroku': ['model_implementation', 'sharepoint_confirmation', ''],
        'przypisany': ['jan.kowalski', '', ''],
        'termin_wykonania': ['2024-12-15', '2024-12-20', ''],
        'priorytet': ['high', 'normal', 'urgent'],
        'status': ['pending', 'in_progress', 'pending'],
        'szacowany_czas': [480, 120, 60],
        'opis': ['Stworzenie modelu CAD', 'Weryfikacja dokumentów', 'Testy końcowe'],
        'notatki': ['', 'Do sprawdzenia z klientem', '']
    }
    df = pd.DataFrame(template_data)
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Szablon_Zadań', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Szablon_Zadań']
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        column_widths = {'A': 25, 'B': 20, 'C': 20, 'D': 15, 'E': 18,
                        'F': 12, 'G': 12, 'H': 15, 'I': 35, 'J': 25}
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, max_col=worksheet.max_column):
            for cell in row:
                cell.border = border
    output.seek(0)
    return output
